#!/usr/bin/env python3
"""
Enhanced Chinese Radicals data pipeline.
Merges Unihan, CC-CEDICT (phrases), and CHISE IDS (decomposition) data.
"""

import json
import zipfile
import io
import re
from pathlib import Path
from collections import defaultdict

# Paths
UNIHAN_PATH = Path("unihan.zip")
CEDICT_PATH = Path("cedict.txt")
CHISE_IDS_DIR = Path("chise-ids-master")
SUBTLEX_PATH = Path("SUBTLEX-CH-WF.xlsx")
SUBTLEX_CHR_PATH = Path("SUBTLEX-CH-CHR.xlsx")
OUTPUT_DIR = Path("data")

# 214 Kangxi Radicals with pinyin and meaning
KANGXI_RADICALS = {
    1: ("一", "yī", "one"),
    2: ("丨", "gǔn", "line"),
    3: ("丶", "zhǔ", "dot"),
    4: ("丿", "piě", "slash"),
    5: ("乙", "yǐ", "second"),
    6: ("亅", "jué", "hook"),
    7: ("二", "èr", "two"),
    8: ("亠", "tóu", "lid"),
    9: ("人", "rén", "person"),
    10: ("儿", "ér", "legs"),
    11: ("入", "rù", "enter"),
    12: ("八", "bā", "eight"),
    13: ("冂", "jiōng", "down box"),
    14: ("冖", "mì", "cover"),
    15: ("冫", "bīng", "ice"),
    16: ("几", "jī", "table"),
    17: ("凵", "kǎn", "open box"),
    18: ("刀", "dāo", "knife"),
    19: ("力", "lì", "power"),
    20: ("勹", "bāo", "wrap"),
    21: ("匕", "bǐ", "spoon"),
    22: ("匚", "fāng", "box"),
    23: ("匸", "xì", "hiding"),
    24: ("十", "shí", "ten"),
    25: ("卜", "bǔ", "divination"),
    26: ("卩", "jié", "seal"),
    27: ("厂", "hàn", "cliff"),
    28: ("厶", "sī", "private"),
    29: ("又", "yòu", "again"),
    30: ("口", "kǒu", "mouth"),
    31: ("囗", "wéi", "enclosure"),
    32: ("土", "tǔ", "earth"),
    33: ("士", "shì", "scholar"),
    34: ("夂", "zhǐ", "go"),
    35: ("夊", "suī", "go slowly"),
    36: ("夕", "xī", "evening"),
    37: ("大", "dà", "big"),
    38: ("女", "nǚ", "woman"),
    39: ("子", "zǐ", "child"),
    40: ("宀", "mián", "roof"),
    41: ("寸", "cùn", "inch"),
    42: ("小", "xiǎo", "small"),
    43: ("尢", "wāng", "lame"),
    44: ("尸", "shī", "corpse"),
    45: ("屮", "chè", "sprout"),
    46: ("山", "shān", "mountain"),
    47: ("巛", "chuān", "river"),
    48: ("工", "gōng", "work"),
    49: ("己", "jǐ", "oneself"),
    50: ("巾", "jīn", "cloth"),
    51: ("干", "gān", "dry"),
    52: ("幺", "yāo", "tiny"),
    53: ("广", "guǎng", "shelter"),
    54: ("廴", "yǐn", "stride"),
    55: ("廾", "gǒng", "hands"),
    56: ("弋", "yì", "shoot"),
    57: ("弓", "gōng", "bow"),
    58: ("彐", "jì", "snout"),
    59: ("彡", "shān", "hair"),
    60: ("彳", "chì", "step"),
    61: ("心", "xīn", "heart"),
    62: ("戈", "gē", "halberd"),
    63: ("戶", "hù", "door"),
    64: ("手", "shǒu", "hand"),
    65: ("支", "zhī", "branch"),
    66: ("攴", "pū", "strike"),
    67: ("文", "wén", "script"),
    68: ("斗", "dǒu", "dipper"),
    69: ("斤", "jīn", "axe"),
    70: ("方", "fāng", "square"),
    71: ("无", "wú", "not"),
    72: ("日", "rì", "sun"),
    73: ("曰", "yuē", "say"),
    74: ("月", "yuè", "moon"),
    75: ("木", "mù", "tree"),
    76: ("欠", "qiàn", "lack"),
    77: ("止", "zhǐ", "stop"),
    78: ("歹", "dǎi", "death"),
    79: ("殳", "shū", "weapon"),
    80: ("毋", "wú", "do not"),
    81: ("比", "bǐ", "compare"),
    82: ("毛", "máo", "fur"),
    83: ("氏", "shì", "clan"),
    84: ("气", "qì", "steam"),
    85: ("水", "shuǐ", "water"),
    86: ("火", "huǒ", "fire"),
    87: ("爪", "zhǎo", "claw"),
    88: ("父", "fù", "father"),
    89: ("爻", "yáo", "lines"),
    90: ("爿", "qiáng", "half tree"),
    91: ("片", "piàn", "slice"),
    92: ("牙", "yá", "fang"),
    93: ("牛", "niú", "cow"),
    94: ("犬", "quǎn", "dog"),
    95: ("玄", "xuán", "dark"),
    96: ("玉", "yù", "jade"),
    97: ("瓜", "guā", "melon"),
    98: ("瓦", "wǎ", "tile"),
    99: ("甘", "gān", "sweet"),
    100: ("生", "shēng", "life"),
    101: ("用", "yòng", "use"),
    102: ("田", "tián", "field"),
    103: ("疋", "pǐ", "bolt of cloth"),
    104: ("疒", "nè", "illness"),
    105: ("癶", "bō", "footsteps"),
    106: ("白", "bái", "white"),
    107: ("皮", "pí", "skin"),
    108: ("皿", "mǐn", "dish"),
    109: ("目", "mù", "eye"),
    110: ("矛", "máo", "spear"),
    111: ("矢", "shǐ", "arrow"),
    112: ("石", "shí", "stone"),
    113: ("示", "shì", "spirit"),
    114: ("禸", "róu", "track"),
    115: ("禾", "hé", "grain"),
    116: ("穴", "xué", "cave"),
    117: ("立", "lì", "stand"),
    118: ("竹", "zhú", "bamboo"),
    119: ("米", "mǐ", "rice"),
    120: ("糸", "mì", "silk"),
    121: ("缶", "fǒu", "jar"),
    122: ("网", "wǎng", "net"),
    123: ("羊", "yáng", "sheep"),
    124: ("羽", "yǔ", "feather"),
    125: ("老", "lǎo", "old"),
    126: ("而", "ér", "and"),
    127: ("耒", "lěi", "plow"),
    128: ("耳", "ěr", "ear"),
    129: ("聿", "yù", "brush"),
    130: ("肉", "ròu", "meat"),
    131: ("臣", "chén", "minister"),
    132: ("自", "zì", "self"),
    133: ("至", "zhì", "arrive"),
    134: ("臼", "jiù", "mortar"),
    135: ("舌", "shé", "tongue"),
    136: ("舛", "chuǎn", "oppose"),
    137: ("舟", "zhōu", "boat"),
    138: ("艮", "gèn", "stop"),
    139: ("色", "sè", "color"),
    140: ("艸", "cǎo", "grass"),
    141: ("虍", "hū", "tiger"),
    142: ("虫", "chóng", "insect"),
    143: ("血", "xuè", "blood"),
    144: ("行", "xíng", "walk"),
    145: ("衣", "yī", "clothes"),
    146: ("襾", "yà", "west"),
    147: ("見", "jiàn", "see"),
    148: ("角", "jiǎo", "horn"),
    149: ("言", "yán", "speech"),
    150: ("谷", "gǔ", "valley"),
    151: ("豆", "dòu", "bean"),
    152: ("豕", "shǐ", "pig"),
    153: ("豸", "zhì", "badger"),
    154: ("貝", "bèi", "shell"),
    155: ("赤", "chì", "red"),
    156: ("走", "zǒu", "run"),
    157: ("足", "zú", "foot"),
    158: ("身", "shēn", "body"),
    159: ("車", "chē", "cart"),
    160: ("辛", "xīn", "bitter"),
    161: ("辰", "chén", "morning"),
    162: ("辵", "chuò", "walk"),
    163: ("邑", "yì", "city"),
    164: ("酉", "yǒu", "wine"),
    165: ("釆", "biàn", "distinguish"),
    166: ("里", "lǐ", "village"),
    167: ("金", "jīn", "gold"),
    168: ("長", "cháng", "long"),
    169: ("門", "mén", "gate"),
    170: ("阜", "fù", "mound"),
    171: ("隶", "lì", "slave"),
    172: ("隹", "zhuī", "bird"),
    173: ("雨", "yǔ", "rain"),
    174: ("青", "qīng", "blue"),
    175: ("非", "fēi", "wrong"),
    176: ("面", "miàn", "face"),
    177: ("革", "gé", "leather"),
    178: ("韋", "wéi", "tanned"),
    179: ("韭", "jiǔ", "leek"),
    180: ("音", "yīn", "sound"),
    181: ("頁", "yè", "leaf"),
    182: ("風", "fēng", "wind"),
    183: ("飛", "fēi", "fly"),
    184: ("食", "shí", "eat"),
    185: ("首", "shǒu", "head"),
    186: ("香", "xiāng", "fragrance"),
    187: ("馬", "mǎ", "horse"),
    188: ("骨", "gǔ", "bone"),
    189: ("高", "gāo", "tall"),
    190: ("髟", "biāo", "hair"),
    191: ("鬥", "dòu", "fight"),
    192: ("鬯", "chàng", "herbs"),
    193: ("鬲", "lì", "cauldron"),
    194: ("鬼", "guǐ", "ghost"),
    195: ("魚", "yú", "fish"),
    196: ("鳥", "niǎo", "bird"),
    197: ("鹵", "lǔ", "salt"),
    198: ("鹿", "lù", "deer"),
    199: ("麥", "mài", "wheat"),
    200: ("麻", "má", "hemp"),
    201: ("黃", "huáng", "yellow"),
    202: ("黍", "shǔ", "millet"),
    203: ("黑", "hēi", "black"),
    204: ("黹", "zhǐ", "embroidery"),
    205: ("黽", "mǐn", "frog"),
    206: ("鼎", "dǐng", "tripod"),
    207: ("鼓", "gǔ", "drum"),
    208: ("鼠", "shǔ", "rat"),
    209: ("鼻", "bí", "nose"),
    210: ("齊", "qí", "even"),
    211: ("齒", "chǐ", "tooth"),
    212: ("龍", "lóng", "dragon"),
    213: ("龜", "guī", "turtle"),
    214: ("龠", "yuè", "flute"),
}


def parse_unihan_readings(zip_data: bytes) -> dict:
    """Parse Unihan_Readings.txt for pinyin and definitions."""
    readings = {}
    
    with zipfile.ZipFile(io.BytesIO(zip_data)) as zf:
        with zf.open("Unihan_Readings.txt") as f:
            for line in f:
                line = line.decode("utf-8").strip()
                if not line or line.startswith("#"):
                    continue
                
                parts = line.split("\t")
                if len(parts) < 3:
                    continue
                
                codepoint, field, value = parts[0], parts[1], parts[2]
                
                if codepoint not in readings:
                    readings[codepoint] = {}
                
                if field == "kMandarin":
                    readings[codepoint]["pinyin"] = value.split()[0].lower()
                elif field == "kDefinition":
                    readings[codepoint]["definition"] = value
    
    return readings


def parse_unihan_radical_stroke(zip_data: bytes) -> dict:
    """Parse Unihan_IRGSources.txt for kRSUnicode (radical + strokes)."""
    radical_data = {}
    
    with zipfile.ZipFile(io.BytesIO(zip_data)) as zf:
        with zf.open("Unihan_IRGSources.txt") as f:
            for line in f:
                line = line.decode("utf-8").strip()
                if not line or line.startswith("#"):
                    continue
                
                parts = line.split("\t")
                if len(parts) < 3:
                    continue
                
                codepoint, field, value = parts[0], parts[1], parts[2]
                
                if field == "kRSUnicode":
                    rs = value.split()[0]
                    rs = rs.replace("'", "")
                    try:
                        radical_str, strokes_str = rs.split(".")
                        radical_data[codepoint] = {
                            "radical": int(radical_str),
                            "strokes": int(strokes_str)
                        }
                    except ValueError:
                        continue
    
    return radical_data


def parse_unihan_grade_level(zip_data: bytes) -> dict:
    """Parse Unihan_DictionaryLikeData.txt for kGradeLevel."""
    grade_data = {}
    
    with zipfile.ZipFile(io.BytesIO(zip_data)) as zf:
        with zf.open("Unihan_DictionaryLikeData.txt") as f:
            for line in f:
                line = line.decode("utf-8").strip()
                if not line or line.startswith("#"):
                    continue
                
                parts = line.split("\t")
                if len(parts) < 3:
                    continue
                
                codepoint, field, value = parts[0], parts[1], parts[2]
                
                if field == "kGradeLevel":
                    try:
                        grade_data[codepoint] = {"gradeLevel": int(value)}
                    except ValueError:
                        continue
    
    return grade_data


def parse_variants(zip_data: bytes) -> tuple:
    """
    Parse Unihan_Variants.txt for simplified/traditional variants.
    Returns: (simp_to_trad, trad_to_simp) mappings
    """
    simp_to_trad = {}  # Simplified char -> Traditional char
    trad_to_simp = {}  # Traditional char -> Simplified char
    
    with zipfile.ZipFile(io.BytesIO(zip_data)) as zf:
        with zf.open("Unihan_Variants.txt") as f:
            for line in f:
                line = line.decode("utf-8").strip()
                if not line or line.startswith("#"):
                    continue
                
                parts = line.split("\t")
                if len(parts) < 3:
                    continue
                
                codepoint, field, value = parts[0], parts[1], parts[2]
                char = chr(int(codepoint[2:], 16))
                
                if field == "kTraditionalVariant":
                    # This char is simplified, value is traditional
                    # Value might have multiple codepoints - skip self-references, take first different one
                    trad_cps = [v.strip() for v in value.split() if v.strip().startswith("U+")]
                    for tcp in trad_cps:
                        trad_char = chr(int(tcp[2:].split("<")[0], 16))
                        if trad_char != char:  # Skip self-reference
                            simp_to_trad[char] = trad_char
                            break
                
                elif field == "kSimplifiedVariant":
                    # This char is traditional, value is simplified
                    # Value might have multiple codepoints - skip self-references, take first different one
                    simp_cps = [v.strip() for v in value.split() if v.strip().startswith("U+")]
                    for scp in simp_cps:
                        simp_char = chr(int(scp[2:].split("<")[0], 16))
                        if simp_char != char:  # Skip self-reference
                            trad_to_simp[char] = simp_char
                            break
    
    return simp_to_trad, trad_to_simp


def parse_cedict(path: Path) -> dict:
    """
    Parse CC-CEDICT dictionary.
    Returns: {
        "by_char": {char: [words containing char]},
        "words": {simplified: {pinyin, definition, traditional}}
    }
    """
    words = {}
    by_char = defaultdict(list)
    
    with open(path, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#"):
                continue
            
            # Format: 傳統 传统 [chuan2 tong3] /tradition/definition.../
            match = re.match(r"^(\S+)\s+(\S+)\s+\[([^\]]+)\]\s+/(.+)/$", line)
            if not match:
                continue
            
            traditional, simplified, pinyin, definitions = match.groups()
            
            # Store word
            words[simplified] = {
                "traditional": traditional,
                "pinyin": pinyin,
                "definition": definitions.split("/")[0]  # First definition
            }
            
            # Index by character
            for char in simplified:
                if simplified not in by_char[char]:
                    by_char[char].append(simplified)
    
    return {"words": words, "by_char": dict(by_char)}


def parse_subtlex(path: Path) -> dict:
    """
    Parse SUBTLEX-CH-WF.xlsx for word frequency.
    Returns: {word: frequency_rank} (1 = most common)
    """
    try:
        import openpyxl
    except ImportError:
        print("  Warning: openpyxl not installed, skipping frequency data")
        return {}
    
    freq_data = {}
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.active
    
    rank = 0
    for row in ws.iter_rows(values_only=True):
        word = row[0]
        if word and isinstance(word, str) and len(word) >= 1:
            # Skip headers
            if word in ("Word", "Total word count: 33,546,516", "Context number: 6,243"):
                continue
            rank += 1
            freq_data[word] = rank
    
    wb.close()
    return freq_data


def parse_subtlex_char(path: Path) -> dict:
    """
    Parse SUBTLEX-CH-CHR.xlsx for character frequency.
    Returns: {char: frequency_rank} (1 = most common)
    """
    try:
        import openpyxl
    except ImportError:
        print("  Warning: openpyxl not installed, skipping character frequency")
        return {}
    
    freq_data = {}
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.active
    
    rank = 0
    for row in ws.iter_rows(values_only=True):
        char = row[0]
        if char and isinstance(char, str) and len(char) == 1:
            # Skip headers
            if char == "C":  # "Character" header
                continue
            rank += 1
            freq_data[char] = rank
    
    wb.close()
    return freq_data


def parse_chise_ids(ids_dir: Path) -> dict:
    """
    Parse CHISE IDS files for character decomposition.
    Returns: {char: {"ids": "⿱宀女", "components": ["宀", "女"]}}
    """
    ids_data = {}
    
    # IDS structural operators
    ids_operators = "⿰⿱⿲⿳⿴⿵⿶⿷⿸⿹⿺⿻"
    
    # Pattern to match CHISE entity references like &CDP-8B7D; or &M-12345;
    entity_pattern = re.compile(r'&[A-Za-z0-9-]+;')
    
    # Common CDP entity references mapped to their Unicode equivalents
    # These are character components that have Unicode codepoints
    cdp_to_unicode = {
        "&CDP-8BCE;": "龰",  # U+9FB0 - bottom of 是,定,走,足
        "&CDP-8B7D;": "⺀",  # top two dots
        "&CDP-89AE;": "龶",  # U+9FB6 - top of 青,靑
        "&CDP-8BF1;": "龷",  # U+9FB7 - top of 共
        "&CDP-8CC6;": "⺁",  # radical variant
        "&CDP-85D5;": "冂",  # cover radical
        "&CDP-89CE;": "⺤",  # claw variant
        "&CDP-8B68;": "止",  # stop radical
        "&CDP-87B5;": "㔾",  # U+353E
        "&CDP-8B42;": "癶",  # U+7676 - footsteps
        "&CDP-89A6;": "⺊",  # divination radical
        "&CDP-8DE2;": "龹",  # U+9FB9
        "&CDP-89AB;": "耂",  # U+8002 - old radical top
        "&CDP-8CB5;": "⺩",  # king radical variant
        "&CDP-8B5E;": "昜",  # U+661C - bright
    }
    
    # Parse main files for common CJK characters
    ids_files = [
        "IDS-UCS-Basic.txt",
        "IDS-UCS-Ext-A.txt",
    ]
    
    for filename in ids_files:
        filepath = ids_dir / filename
        if not filepath.exists():
            continue
        
        with open(filepath, "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line or line.startswith(";"):
                    continue
                
                parts = line.split("\t")
                if len(parts) < 3:
                    continue
                
                codepoint, char, ids_raw = parts[0], parts[1], parts[2]
                
                # First, replace known CDP entity references with Unicode equivalents
                ids_processed = ids_raw
                for entity, unicode_char in cdp_to_unicode.items():
                    ids_processed = ids_processed.replace(entity, unicode_char)
                
                # Handle remaining entity references (incomplete decomposition)
                if entity_pattern.search(ids_processed):
                    # Remove remaining entity references
                    ids_clean = entity_pattern.sub('', ids_processed)
                    # If only operators remain or it's too short, skip
                    non_operator_chars = [c for c in ids_clean if c not in ids_operators]
                    if len(non_operator_chars) < 2:
                        continue
                    ids = ids_clean
                else:
                    ids = ids_processed
                
                # Extract components (filter out IDS operators and the character itself)
                components = []
                for c in ids:
                    if c not in ids_operators and c != char and len(c) == 1:
                        # Only include valid CJK characters (Basic CJK range)
                        code = ord(c)
                        if (0x4E00 <= code <= 0x9FFF or  # CJK Unified Ideographs
                            0x3400 <= code <= 0x4DBF or  # CJK Extension A
                            0x2E80 <= code <= 0x2EFF or  # CJK Radicals Supplement
                            0x2F00 <= code <= 0x2FDF):   # Kangxi Radicals
                            components.append(c)
                
                # Only store if we have valid decomposition
                if ids != char and len(components) >= 1:
                    ids_data[char] = {
                        "ids": ids,
                        "components": components
                    }
    
    return ids_data


def build_appears_in(ids_data: dict) -> dict:
    """
    Build reverse lookup: for each component, find chars it appears in.
    Returns: {component: [char1, char2, ...]}
    """
    appears_in = defaultdict(list)
    
    for char, data in ids_data.items():
        for component in data.get("components", []):
            if component and len(component) == 1:
                appears_in[component].append(char)
    
    return dict(appears_in)


def codepoint_to_char(cp: str) -> str:
    """Convert U+XXXX to actual character."""
    return chr(int(cp[2:], 16))


def build_radicals_json(readings: dict, radical_data: dict, grade_data: dict,
                        cedict: dict, ids_data: dict, appears_in: dict,
                        word_freq: dict, char_freq: dict,
                        simp_to_trad: dict, trad_to_simp: dict) -> dict:
    """Build the final unified JSON structure."""
    
    # Build radicals section
    radicals = {}
    for num, (char, pinyin, meaning) in KANGXI_RADICALS.items():
        radicals[str(num)] = {
            "char": char,
            "pinyin": pinyin,
            "meaning": meaning,
            "characters": []
        }
    
    # Build characters section
    characters = {}
    
    for codepoint, rs in radical_data.items():
        radical_num = rs["radical"]
        if radical_num not in KANGXI_RADICALS:
            continue
        
        char = codepoint_to_char(codepoint)
        char_readings = readings.get(codepoint, {})
        char_grade = grade_data.get(codepoint, {})
        
        # Only include characters with definitions
        if "definition" not in char_readings:
            continue
        
        # Get IDS decomposition
        ids_info = ids_data.get(char, {})
        
        # Get words containing this character (limit to 30, sorted by frequency)
        char_words = cedict["by_char"].get(char, [])
        # Sort by frequency (lower rank = more common), then alphabetically
        char_words = sorted(char_words, key=lambda w: (word_freq.get(w, 999999), w))[:30]
        
        # Get characters this appears in as component (limit to 20)
        derived_chars = appears_in.get(char, [])
        
        char_data = {
            "radical": str(radical_num),
            "strokes": rs["strokes"],
            "pinyin": char_readings.get("pinyin", ""),
            "definition": char_readings["definition"],
            "gradeLevel": char_grade.get("gradeLevel", 0),
            "charFrequency": char_freq.get(char, 0),  # 0 = not ranked
            "ids": ids_info.get("ids", ""),
            "components": ids_info.get("components", []),
            "words": char_words,
            "appearsIn": derived_chars,
            "traditional": simp_to_trad.get(char),  # If simplified, link to traditional
            "simplified": trad_to_simp.get(char),   # If traditional, link to simplified
        }
        
        characters[char] = char_data
        radicals[str(radical_num)]["characters"].append(char)
    
    # Sort characters within each radical by grade level then strokes
    for radical in radicals.values():
        radical["characters"].sort(
            key=lambda c: (
                0 if characters[c]["gradeLevel"] > 0 else 1,
                characters[c]["gradeLevel"] if characters[c]["gradeLevel"] > 0 else 99,
                characters[c]["strokes"]
            )
        )
    
    # Second pass: Add component characters that aren't full characters but have Unihan readings
    # These are characters used in IDS decomposition but not in main character set
    all_components = set()
    for char_data in characters.values():
        for comp in char_data.get("components", []):
            if comp not in characters:
                all_components.add(comp)
    
    added_components = 0
    for comp in all_components:
        # Try to find this component in Unihan readings
        comp_codepoint = f"U+{ord(comp):04X}"
        comp_readings = readings.get(comp_codepoint, {})
        
        if comp_readings.get("pinyin"):
            # This component has Unihan data - add it
            comp_radical = radical_data.get(comp_codepoint, {})
            characters[comp] = {
                "radical": str(comp_radical.get("radical", 0)),
                "strokes": comp_radical.get("strokes", 0),
                "pinyin": comp_readings.get("pinyin", ""),
                "definition": comp_readings.get("definition", "component"),
                "gradeLevel": 0,
                "charFrequency": char_freq.get(comp, 0),
                "ids": ids_data.get(comp, {}).get("ids", ""),
                "components": ids_data.get(comp, {}).get("components", []),
                "words": [],
                "appearsIn": appears_in.get(comp, []),
                "traditional": simp_to_trad.get(comp),
                "simplified": trad_to_simp.get(comp),
            }
            added_components += 1
    
    print(f"  Added {added_components} component characters with Unihan readings")

    # Post-process: inherit metadata between simplified/traditional variants
    inherited_grades = 0
    inherited_freq = 0
    for char, char_data in characters.items():
        # Simplified inherits grade from traditional
        if char_data["gradeLevel"] == 0 and char_data.get("traditional"):
            trad = char_data["traditional"]
            if trad in characters and characters[trad]["gradeLevel"] > 0:
                char_data["gradeLevel"] = characters[trad]["gradeLevel"]
                inherited_grades += 1
        
        # Traditional inherits frequency from simplified
        if char_data["charFrequency"] == 0 and char_data.get("simplified"):
            simp = char_data["simplified"]
            if simp in characters and characters[simp]["charFrequency"] > 0:
                char_data["charFrequency"] = characters[simp]["charFrequency"]
                inherited_freq += 1
    
    print(f"  Inherited {inherited_grades} grades (simp←trad), {inherited_freq} frequencies (trad←simp)")
    
    # Add words dictionary - include all words that are referenced by characters
    referenced_words = set()
    for char_data in characters.values():
        for word in char_data.get("words", []):
            referenced_words.add(word)
    
    words_subset = {}
    for word in referenced_words:
        if word in cedict["words"]:
            words_subset[word] = {
                **cedict["words"][word],
                "frequency": word_freq.get(word, 0)
            }
    
    print(f"  Words referenced by characters: {len(referenced_words)}, with data: {len(words_subset)}")
    
    return {
        "radicals": radicals,
        "characters": characters,
        "words": words_subset
    }


def main():
    # Load Unihan
    print("Loading Unihan database...")
    with open(UNIHAN_PATH, "rb") as f:
        zip_data = f.read()
    print(f"  Loaded {len(zip_data) / 1024 / 1024:.1f} MB")
    
    # Parse Unihan
    print("Parsing Unihan_Readings.txt...")
    readings = parse_unihan_readings(zip_data)
    print(f"  Found {len(readings)} characters with readings")
    
    print("Parsing Unihan_IRGSources.txt...")
    radical_data = parse_unihan_radical_stroke(zip_data)
    print(f"  Found {len(radical_data)} characters with radical data")
    
    print("Parsing Unihan_DictionaryLikeData.txt...")
    grade_data = parse_unihan_grade_level(zip_data)
    print(f"  Found {len(grade_data)} characters with grade level data")
    
    # Parse CC-CEDICT
    print("Parsing CC-CEDICT...")
    cedict = parse_cedict(CEDICT_PATH)
    print(f"  Found {len(cedict['words'])} words")
    print(f"  Indexed {len(cedict['by_char'])} unique characters")
    
    # Parse CHISE IDS
    print("Parsing CHISE IDS...")
    ids_data = parse_chise_ids(CHISE_IDS_DIR)
    print(f"  Found {len(ids_data)} characters with decomposition data")
    
    # Build reverse lookup
    print("Building 'appears in' index...")
    appears_in = build_appears_in(ids_data)
    print(f"  Indexed {len(appears_in)} components")
    
    # Parse SUBTLEX frequency
    print("Parsing SUBTLEX-CH word frequency...")
    if SUBTLEX_PATH.exists():
        word_freq = parse_subtlex(SUBTLEX_PATH)
        print(f"  Found {len(word_freq)} words with frequency data")
    else:
        print("  SUBTLEX file not found, skipping frequency data")
        word_freq = {}
    
    # Parse SUBTLEX character frequency
    print("Parsing SUBTLEX-CH character frequency...")
    if SUBTLEX_CHR_PATH.exists():
        char_freq = parse_subtlex_char(SUBTLEX_CHR_PATH)
        print(f"  Found {len(char_freq)} characters with frequency data")
    else:
        print("  SUBTLEX-CHR file not found, skipping character frequency")
        char_freq = {}
    
    # Parse simplified/traditional variants
    print("Parsing character variants...")
    simp_to_trad, trad_to_simp = parse_variants(zip_data)
    print(f"  Found {len(simp_to_trad)} simplified→traditional, {len(trad_to_simp)} traditional→simplified")
    
    # Build final JSON
    print("Building unified radicals.json...")
    result = build_radicals_json(readings, radical_data, grade_data, 
                                  cedict, ids_data, appears_in, word_freq, char_freq,
                                  simp_to_trad, trad_to_simp)
    
    # Statistics
    total_chars = len(result["characters"])
    chars_with_words = sum(1 for c in result["characters"].values() if c["words"])
    chars_with_ids = sum(1 for c in result["characters"].values() if c["ids"])
    
    print(f"  Total characters: {total_chars}")
    print(f"  Characters with words: {chars_with_words}")
    print(f"  Characters with IDS: {chars_with_ids}")
    
    # Write output
    OUTPUT_DIR.mkdir(exist_ok=True)
    output_path = OUTPUT_DIR / "radicals.json"
    
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(result, f, ensure_ascii=False, indent=2)
    
    print(f"\nOutput written to {output_path}")
    print(f"File size: {output_path.stat().st_size / 1024 / 1024:.1f} MB")


if __name__ == "__main__":
    main()
